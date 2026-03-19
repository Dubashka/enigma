import io
import pytest
import openpyxl


@pytest.fixture
def sample_xlsx_bytes():
    """3-sheet workbook: Лист1 (3 rows, 3 cols), Лист2 (2 rows, 2 cols), Лист3 (empty)."""
    wb = openpyxl.Workbook()
    # Sheet 1: Лист1
    ws1 = wb.active
    ws1.title = "Лист1"
    ws1.append(["Название", "Количество", "Цена"])
    ws1.append(["Товар А", 10, 100.0])
    ws1.append(["Товар Б", 20, 200.0])
    ws1.append(["Товар В", 30, 300.0])
    # Sheet 2: Лист2
    ws2 = wb.create_sheet("Лист2")
    ws2.append(["Код", "Значение"])
    ws2.append(["001", 42])
    ws2.append(["002", 99])
    # Sheet 3: Лист3 — empty (no data rows)
    wb.create_sheet("Лист3")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture
def sample_csv_bytes_utf8():
    """CSV with Cyrillic headers, semicolon separator, UTF-8 encoding."""
    content = "Название;Количество;Цена\nТовар А;10;100.0\nТовар Б;20;200.0\nТовар В;30;300.0\n"
    return content.encode("utf-8")


@pytest.fixture
def sample_csv_bytes_cp1251():
    """Same CSV content encoded as cp1251."""
    content = "Название;Количество;Цена\nТовар А;10;100.0\nТовар Б;20;200.0\nТовар В;30;300.0\n"
    return content.encode("cp1251")


@pytest.fixture
def sample_csv_comma():
    """CSV with comma separator, UTF-8, Cyrillic headers."""
    content = "Название,Количество,Цена\nТовар А,10,100.0\nТовар Б,20,200.0\nТовар В,30,300.0\n"
    return content.encode("utf-8")


@pytest.fixture
def corrupt_bytes():
    """Random invalid bytes that cannot be parsed as xlsx or csv."""
    return b"not a valid file at all" * 5
