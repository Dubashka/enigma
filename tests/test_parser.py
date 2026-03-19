"""Tests for core/parser.py — covers all LOAD-01, LOAD-02, LOAD-03, UI-02 requirements."""
import io
import pytest
import openpyxl
from dataclasses import dataclass

from core.parser import parse_upload


@dataclass
class FakeUploadedFile:
    name: str
    _data: bytes

    def read(self) -> bytes:
        return self._data


# ---------------------------------------------------------------------------
# LOAD-01: Excel multi-sheet support
# ---------------------------------------------------------------------------

def test_parse_excel_multisheet(sample_xlsx_bytes):
    """xlsx with 3 sheets (one empty) returns dict with 2 non-empty sheet keys."""
    f = FakeUploadedFile(name="test.xlsx", _data=sample_xlsx_bytes)
    result = parse_upload(f)
    assert isinstance(result, dict)
    # Лист3 is empty — should be filtered
    assert len(result) == 2
    assert "Лист1" in result
    assert "Лист2" in result
    import pandas as pd
    assert isinstance(result["Лист1"], pd.DataFrame)
    assert isinstance(result["Лист2"], pd.DataFrame)


def test_empty_sheets_filtered(sample_xlsx_bytes):
    """xlsx with 1 populated + 1 empty sheet returns dict with 1 key only when all others empty."""
    # Build a workbook with 1 populated + 1 empty sheet
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Данные"
    ws1.append(["A", "B"])
    ws1.append([1, 2])
    wb.create_sheet("Пустой")  # empty sheet
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    f = FakeUploadedFile(name="test.xlsx", _data=buf.read())
    result = parse_upload(f)
    assert len(result) == 1
    assert "Данные" in result
    assert "Пустой" not in result


def test_all_sheets_empty():
    """xlsx where all sheets are empty raises ValueError with Russian message."""
    wb = openpyxl.Workbook()
    # Default sheet is empty, add another empty one
    wb.active.title = "Лист1"
    wb.create_sheet("Лист2")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    f = FakeUploadedFile(name="empty.xlsx", _data=buf.read())
    with pytest.raises(ValueError, match="Файл не содержит данных"):
        parse_upload(f)


# ---------------------------------------------------------------------------
# LOAD-02: CSV parsing
# ---------------------------------------------------------------------------

def test_parse_csv_utf8(sample_csv_bytes_utf8):
    """CSV with UTF-8 content returns {'Лист1': DataFrame}."""
    f = FakeUploadedFile(name="data.csv", _data=sample_csv_bytes_utf8)
    result = parse_upload(f)
    assert isinstance(result, dict)
    assert "Лист1" in result
    assert len(result) == 1
    import pandas as pd
    assert isinstance(result["Лист1"], pd.DataFrame)
    assert len(result["Лист1"]) == 3


def test_parse_csv_cp1251(sample_csv_bytes_cp1251):
    """CSV encoded in cp1251 with Cyrillic headers returns correct column names (not garbled)."""
    f = FakeUploadedFile(name="data.csv", _data=sample_csv_bytes_cp1251)
    result = parse_upload(f)
    assert "Лист1" in result
    df = result["Лист1"]
    columns = list(df.columns)
    assert "Название" in columns
    assert "Количество" in columns
    assert "Цена" in columns


def test_parse_csv_semicolon(sample_csv_bytes_utf8):
    """CSV with semicolon separator parses correctly (columns are not merged)."""
    f = FakeUploadedFile(name="data.csv", _data=sample_csv_bytes_utf8)
    result = parse_upload(f)
    df = result["Лист1"]
    # Should have 3 columns, not 1 merged column
    assert len(df.columns) == 3


# ---------------------------------------------------------------------------
# LOAD-03: Structure preservation
# ---------------------------------------------------------------------------

def test_column_order_preserved():
    """Columns in returned DataFrame match source order exactly."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    headers = ["Дата", "Контрагент", "Сумма", "Договор"]
    ws.append(headers)
    ws.append(["2024-01-01", "ООО Ромашка", 1000.0, "Д-001"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    f = FakeUploadedFile(name="test.xlsx", _data=buf.read())
    result = parse_upload(f)
    assert list(result["Лист1"].columns) == headers


def test_row_order_preserved():
    """First and last row values match source file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(["Название", "Значение"])
    ws.append(["Первый", 1])
    ws.append(["Средний", 2])
    ws.append(["Последний", 3])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    f = FakeUploadedFile(name="test.xlsx", _data=buf.read())
    result = parse_upload(f)
    df = result["Лист1"]
    assert df.iloc[0]["Название"] == "Первый"
    assert df.iloc[-1]["Название"] == "Последний"


# ---------------------------------------------------------------------------
# Error handling: unsupported format and corrupt files
# ---------------------------------------------------------------------------

def test_unsupported_format(corrupt_bytes):
    """File with .txt extension raises ValueError with Russian message."""
    f = FakeUploadedFile(name="data.txt", _data=corrupt_bytes)
    with pytest.raises(ValueError, match="Поддерживаются только файлы xlsx и csv"):
        parse_upload(f)


def test_corrupt_xlsx(corrupt_bytes):
    """Invalid bytes with .xlsx name raises ValueError with Russian message."""
    f = FakeUploadedFile(name="bad.xlsx", _data=corrupt_bytes)
    with pytest.raises(ValueError, match="Не удалось прочитать файл"):
        parse_upload(f)


def test_corrupt_csv(corrupt_bytes):
    """Completely unparseable bytes with .csv name raises ValueError with Russian message."""
    # Use bytes that cannot be decoded in any encoding
    bad_data = bytes(range(256)) * 3  # raw bytes 0x00-0xFF, invalid in all encodings
    f = FakeUploadedFile(name="bad.csv", _data=bad_data)
    with pytest.raises(ValueError, match="Не удалось прочитать файл"):
        parse_upload(f)
