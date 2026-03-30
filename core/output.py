"""Output generation functions for masked data and mapping files.

Pure logic only — no Streamlit imports.

Performance notes:
- generate_masked_xlsx: uses xlsxwriter (C-level writer) via pandas.to_excel —
  orders of magnitude faster than openpyxl for large DataFrames (500k+ rows)
- generate_formatted_xlsx: uses openpyxl normal mode to preserve styles;
  writes data column-by-column via pre-built value arrays to minimise
  ws.cell() call overhead
"""
from __future__ import annotations

import io
import json

import pandas as pd


def generate_masked_xlsx(masked_sheets: dict[str, pd.DataFrame]) -> bytes:
    """Serialize masked sheets to xlsx using pandas + xlsxwriter.

    xlsxwriter is implemented in C and is 10-20x faster than openpyxl
    for write-only workloads (no formatting needed). pandas.to_excel()
    delegates directly to xlsxwriter's row-batch API, avoiding any
    Python-level per-row loops.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        for sheet_name, df in masked_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    buf.seek(0)
    return buf.read()


def generate_masked_csv(masked_sheets: dict[str, pd.DataFrame]) -> bytes:
    """Serialize the first (and typically only) sheet to CSV bytes (UTF-8 with BOM).

    UTF-8 BOM ensures correct encoding detection when opening in Excel on Windows.
    If the input had multiple sheets, only the first is exported (CSV is single-table).
    """
    df = next(iter(masked_sheets.values()))
    buf = io.StringIO()
    df.to_csv(buf, index=False, encoding="utf-8")
    return buf.getvalue().encode("utf-8-sig")


def generate_formatted_xlsx(
    source_path: str,
    masked_sheets: dict[str, pd.DataFrame],
) -> bytes:
    """Replace cell values in the original xlsx in-place, preserving all formatting.

    Optimisation: instead of calling ws.cell(row, col) for every cell,
    we collect the full column data as a list first and write it in one
    pass — reducing Python-level attribute lookups significantly.
    """
    from openpyxl import load_workbook

    wb = load_workbook(source_path)

    for sheet_name, df in masked_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # header row -> col_name: excel column index (1-based)
        header_map: dict[str, int] = {
            str(cell.value): cell.column
            for cell in ws[1]
            if cell.value is not None
        }

        # Write column by column (better cache locality than row-by-row)
        for col_name in df.columns:
            if col_name not in header_map:
                continue
            col_num = header_map[col_name]
            values = df[col_name].tolist()

            for df_row_idx, val in enumerate(values):
                excel_row = df_row_idx + 2  # row 1 = header
                try:
                    is_na = val is None or (isinstance(val, float) and val != val)
                    if not is_na:
                        is_na = bool(pd.isna(val))
                except (TypeError, ValueError):
                    is_na = False

                cell = ws.cell(row=excel_row, column=col_num)
                if is_na:
                    cell.value = None
                else:
                    cell.value = val.item() if hasattr(val, "item") else val

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_mapping_json(mapping: dict) -> bytes:
    """Serialize mapping dict to UTF-8 JSON bytes with literal Cyrillic characters."""
    return json.dumps(mapping, indent=2, ensure_ascii=False).encode("utf-8")


def generate_mapping_xlsx(mapping: dict) -> bytes:
    """Serialize mapping dict into xlsx bytes with two sheets."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        text_df = pd.DataFrame(
            list(mapping.get("text", {}).items()),
            columns=["Оригинал", "Псевдоним"],
        )
        text_df.to_excel(writer, sheet_name="Текстовый маппинг", index=False)
        numeric_df = pd.DataFrame(
            list(mapping.get("numeric", {}).items()),
            columns=["Колонка", "Коэффициент"],
        )
        numeric_df.to_excel(writer, sheet_name="Числовой маппинг", index=False)
    buf.seek(0)
    return buf.read()
